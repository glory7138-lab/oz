import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import os

base_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(base_dir, "26C17,18 Transmittal2.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=False)
sheet = wb.active

output = []
output.append(f"=== Sheet: {sheet.title} ===")
output.append(f"Max rows: {sheet.max_row}, Max cols: {sheet.max_column}")
output.append("")

output.append("=== MERGED CELLS ===")
for m in sheet.merged_cells.ranges:
    output.append(f"  {m}")
output.append("")

output.append("=== ALL CELLS (row 1-35) ===")
for r in range(1, 36):
    row_data = []
    for c in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        if val is not None and str(val).strip() != "":
            col_letter = get_column_letter(c)
            font = cell.font
            align = cell.alignment
            fill = cell.fill
            
            font_info = f"bold={font.bold}, size={font.size}, name={font.name}" if font else ""
            align_info = f"h={align.horizontal}, v={align.vertical}, wrap={align.wrap_text}" if align else ""
            fill_info = f"fg={fill.fgColor.rgb if fill and fill.fgColor else 'none'}" if fill else ""
            
            row_data.append(f"  [{r},{col_letter}] val='{val}' | {font_info} | {align_info} | {fill_info}")
    
    if row_data:
        output.append(f"--- Row {r} ---")
        output.extend(row_data)

output.append("")
output.append("=== COLUMN WIDTHS ===")
for col in sheet.column_dimensions:
    output.append(f"  Col {col}: width={sheet.column_dimensions[col].width}")

output.append("")
output.append("=== ROW HEIGHTS ===")
for row_num in range(1, 36):
    rd = sheet.row_dimensions.get(row_num)
    if rd:
        output.append(f"  Row {row_num}: height={rd.height}")

out_txt_path = os.path.join(base_dir, "excel_analysis_full.txt")
with open(out_txt_path, "w", encoding="utf-8") as f:
    f.write("\n".join(output))

print("Done! Written to excel_analysis_full.txt")
print("\n".join(output[:80]))
