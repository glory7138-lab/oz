import openpyxl
from openpyxl.utils import get_column_letter
import sys, json

import os

base_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.join(base_dir, "26C17,18 Transmittal2.xlsx")
if not os.path.exists(xlsx_path):
    xlsx_path = r"C:\CPE_DEV\workspace\CPE_APP\war\reports\pl\pla\26C17,18 Transmittal2.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    print(f"Rows: {ws.min_row} - {ws.max_row}")
    print(f"Cols: {ws.min_column} - {ws.max_column}")
    print(f"Merged cells: {[str(m) for m in ws.merged_cells.ranges]}")
    print()
    
    for r in range(ws.min_row, ws.max_row + 1):
        cells = []
        for c in range(ws.min_column, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                bold = "B" if cell.font.bold else " "
                sz = cell.font.size if cell.font.size else "-"
                ha = cell.alignment.horizontal if cell.alignment and cell.alignment.horizontal else "-"
                fname = cell.font.name if cell.font.name else "-"
                col_letter = get_column_letter(c)
                cells.append(f"{col_letter}:[{bold}|{sz}|{ha}|{fname}] {repr(val)}")
        
        if cells:
            print(f"Row {r:3d}: " + " | ".join(cells))
        else:
            print(f"Row {r:3d}: (empty)")
    
    print()
    
    # Column widths
    print("Column widths:")
    for col_idx in range(ws.min_column, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        dim = ws.column_dimensions.get(col_letter)
        width = dim.width if dim and dim.width else "default"
        print(f"  {col_letter}: {width}")
    
    # Row heights
    print("\nRow heights:")
    for r in range(ws.min_row, ws.max_row + 1):
        dim = ws.row_dimensions.get(r)
        height = dim.height if dim and dim.height else "default"
        print(f"  Row {r}: {height}")

print("\nDone!")
