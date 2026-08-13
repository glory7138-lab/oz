"""
Script to rewrite xlsx_parser.py and ozr_generator.py for exact layout reproduction
"""
import os

xlsx_parser_code = '''"""
XLSX 파서 모듈 (개선판)
병합 셀(Merged Cells)과 데이터 바인딩(<DATASET:COLNAME>) 패턴을 완벽히 인식하여 레이아웃을 정확히 추출합니다.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import io
import re

EXCEL_COL_WIDTH_TO_OZ = 7.0
EXCEL_ROW_HEIGHT_TO_OZ = 1.33

def parse_xlsx(filepath) -> dict:
    if isinstance(filepath, bytes):
        file_bytes = io.BytesIO(filepath)
    else:
        with open(filepath, "rb") as f:
            file_bytes = io.BytesIO(f.read())
            
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    ws = wb.active
    
    # 컬럼 폭 계산
    col_widths = []
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        dim = ws.column_dimensions.get(col_letter)
        width = dim.width if dim and dim.width else 8.43
        col_widths.append(width * EXCEL_COL_WIDTH_TO_OZ)
        
    # 누적 X 좌표 (컬럼별)
    col_x_offsets = [0]
    for w in col_widths:
        col_x_offsets.append(col_x_offsets[-1] + w)
        
    # 누적 Y 좌표 (로우별)
    max_row = ws.max_row or 1
    row_heights = []
    for r in range(1, max_row + 1):
        rd = ws.row_dimensions.get(r)
        height = rd.height if rd and rd.height else 15.0
        row_heights.append(height * EXCEL_ROW_HEIGHT_TO_OZ)
        
    row_y_offsets = [0]
    for h in row_heights:
        row_y_offsets.append(row_y_offsets[-1] + h)
        
    # 병합 셀 맵핑 (Merged Cell Range 찾기)
    merged_ranges = ws.merged_cells.ranges
    merged_map = {} # (r, c) -> range
    for mr in merged_ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_map[(r, c)] = mr
                
    cells_data = []
    
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if (r, c) in merged_map:
                mr = merged_map[(r, c)]
                # 병합된 영역의 Top-Left 셀이 아니면 무시 (그리기 중복 방지)
                if r != mr.min_row or c != mr.min_col:
                    continue
                
                # 너비와 높이 계산
                left = col_x_offsets[mr.min_col - 1]
                top = row_y_offsets[mr.min_row - 1]
                width = col_x_offsets[mr.max_col] - left
                height = row_y_offsets[mr.max_row] - top
            else:
                left = col_x_offsets[c - 1]
                top = row_y_offsets[r - 1]
                width = col_x_offsets[c] - left
                height = row_y_offsets[r] - top
                
            cell = ws.cell(row=r, column=c)
            
            # 값이 있거나 테두리가 있거나 배경색이 있으면 유효한 셀로 간주
            val = cell.value
            val_str = str(val).strip() if val is not None else ""
            has_border = _has_border(cell)
            bg_color = _get_fill_color(cell)
            
            if not val_str and not has_border and bg_color == "none":
                continue
                
            cells_data.append({
                "row": r,
                "col": c,
                "text": val_str,
                "left": left,
                "top": top,
                "width": width,
                "height": height,
                "bold": cell.font.bold if cell.font else False,
                "font_size": cell.font.size if cell.font and cell.font.size else 10,
                "h_align": _map_h_align(cell.alignment.horizontal if cell.alignment else "left"),
                "has_border": has_border,
                "bg_color": bg_color
            })
            
    # 영역 분류 (Data Row 찾기: <DATASET:COLNAME> 패턴이 있는 행)
    data_row_num = -1
    for cell in cells_data:
        if re.search(r'<[^:]+:[^>]+>', cell["text"]):
            data_row_num = cell["row"]
            break
            
    if data_row_num == -1:
        # 패턴이 없으면 기존 로직처럼 대략적으로 중간을 바디로 잡거나, 전체를 헤더로
        data_row_num = max_row // 2 if max_row > 2 else 2
        
    # Table Header Rows 식별 (Data Row 바로 위 연속된 테두리 있는 행들)
    table_header_start = data_row_num
    for r in range(data_row_num - 1, 0, -1):
        # 해당 행에 테두리 있는 셀이 하나라도 있는지
        has_borders = any(c["has_border"] for c in cells_data if c["row"] == r)
        if has_borders:
            table_header_start = r
        else:
            break
            
    header_cells = []
    table_header_cells = []
    table_data_cells = []
    footer_cells = []
    
    for c in cells_data:
        r = c["row"]
        if r < table_header_start:
            header_cells.append(c)
        elif table_header_start <= r < data_row_num:
            table_header_cells.append(c)
        elif r == data_row_num:
            table_data_cells.append(c)
        else:
            footer_cells.append(c)
            
    # 높이 계산
    header_height = row_y_offsets[table_header_start - 1] if table_header_start > 1 else 0
    
    # 풋터 좌표 보정 (풋터는 DataBand 바로 밑에서 시작하므로 top을 0부터 다시 시작하게끔 보정)
    footer_start_y = row_y_offsets[data_row_num]
    for c in footer_cells:
        c["top"] -= footer_start_y
    footer_height = row_y_offsets[-1] - footer_start_y if footer_start_y < row_y_offsets[-1] else 20
    
    # DataBand 테이블의 Y 좌표 보정 (Table Header부터 0)
    table_start_y = row_y_offsets[table_header_start - 1]
    for c in table_header_cells + table_data_cells:
        c["top"] -= table_start_y
        
    body_height = row_y_offsets[data_row_num] - table_start_y
    
    result = {
        "sheet_name": ws.title,
        "header": {
            "height": max(header_height, 20),
            "labels": header_cells
        },
        "body": {
            "title_labels": table_header_cells,
            "data_labels": table_data_cells,
            "height": max(body_height, 40)
        },
        "footer": {
            "height": max(footer_height, 20),
            "labels": footer_cells
        },
        "total_width": col_x_offsets[-1]
    }
    
    wb.close()
    return result

def _get_fill_color(cell) -> str:
    try:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
            return str(fill.fgColor.rgb)
    except:
        pass
    return "none"

def _has_border(cell) -> bool:
    try:
        border = cell.border
        if border:
            for side in [border.left, border.right, border.top, border.bottom]:
                if side and side.style and side.style != "none":
                    return True
    except:
        pass
    return False

def _map_h_align(align_str: str) -> str:
    # 0=left, 1=center, 2=right
    mapping = {"left": "0", "center": "1", "right": "2", "general": "0"}
    return mapping.get(str(align_str).lower(), "0")
'''

ozr_generator_code = '''"""
OZR (OZ Report) 파일 생성 모듈 (정확한 좌표 및 데이터 바인딩 적용)
"""

import time
import html
import re

OZR_HEADER = b'OZR\\x07\\x00\\x00\\x00\\x0e\\x00\\x00OZ Report File'

PAPER_WIDTH = 842
PAPER_HEIGHT = 595
LEFT_MARGIN = 19.991
TOP_MARGIN = 15.011
RIGHT_MARGIN = 15.991
BOTTOM_MARGIN = 0
CONTENT_WIDTH = PAPER_WIDTH - LEFT_MARGIN - RIGHT_MARGIN

def escape_xml(text: str) -> str:
    return html.escape(str(text), quote=True)

def _build_version():
    timestamp = int(time.time() * 1000)
    return f'<VERSION VERSION="7.0" DATE="{timestamp}"/>'

def _build_basic_label():
    return (
        '<BASICLABEL WIDTH="100" HEIGHT="20" BGCOLOR="-1" TRANSPARENT="false" '
        'BackgroundAlpha="255" CLIP="false" DRAWLEFT="0.125" DRAWTOP="0.125" '
        'DRAWRIGHT="0.125" DRAWBOTTOM="0.125" DASHLEFT="" DASHTOP="" '
        'DASHRIGHT="" DASHBOTTOM="" DASHOFFSETLEFT="" DASHOFFSETTOP="" '
        'DASHOFFSETRIGHT="" DASHOFFSETBOTTOM="" FRAMECOLORLEFT="-16777216" '
        'FRAMECOLORTOP="-16777216" FRAMECOLORRIGHT="-16777216" '
        'FRAMECOLORBOTTOM="-16777216" FrameDrawingMode="2" RADIUSTOPLEFT="0" '
        'RADIUSTOPRIGHT="0" RADIUSBOTTOMRIGHT="0" RADIUSBOTTOMLEFT="0" '
        'FGCOLOR="-16777216" FONTNAME="맑은 고딕" FONTSIZE="10" FONTSTYLE="0" '
        'STRETCHTYPE="1" HALIGN="0" VALIGN="0" TEXTROTATION="0" '
        'EFFECT="Basic" SPACING="0" FONTSTRETCH="100" WRAPSPACE="0" '
        'LEFTMARGIN="2" TOPMARGIN="0" RIGHTMARGIN="2" BOTTOMMARGIN="0" '
        'AUTOSIZE="false" AUTOFONTSIZE="false" WORDWRAP="false" '
        'WORDWRAPTYPE="1" CRLFTOLF="false" TABCOUNT="4" '
        'USEGRADIENT="False" GRADIENTCOLOR="-1" GRADIENTTYPE="6"/>'
    )

def _build_default_label():
    return '<DEFAULTLABEL TABCOUNT="4"/>'

def _build_static_table_labels(labels: list, prefix: str) -> str:
    """OZTABLELABEL 들을 생성합니다."""
    xml = []
    for idx, lbl in enumerate(labels):
        text = escape_xml(lbl.get("text", ""))
        left = round(lbl.get("left", 0), 3)
        top = round(lbl.get("top", 0), 3)
        width = round(lbl.get("width", 100), 3)
        height = round(lbl.get("height", 20), 3)
        h_align = lbl.get("h_align", "0")
        font_size = lbl.get("font_size", 10)
        font_style = "1" if lbl.get("bold") else "0"
        
        border_attrs = 'DRAWLEFT="1" DRAWTOP="1" DRAWRIGHT="1" DRAWBOTTOM="1" ' if lbl.get("has_border") else ""
        bg_color_attr = f'BGCOLOR="14803425" ' if lbl.get("bg_color") != "none" and lbl.get("bg_color") else ""
        
        xml.append(
            f'<OZTABLELABEL NAME="{prefix}{idx+1}" LEFT="{left}" TOP="{top}" '
            f'WIDTH="{width}" HEIGHT="{height}" '
            f'DASHOFFSETLEFT="0" DASHOFFSETTOP="0" DASHOFFSETRIGHT="0" DASHOFFSETBOTTOM="0" '
            f'{border_attrs}{bg_color_attr}'
            f'FONTSIZE="{font_size}" FONTSTYLE="{font_style}" HALIGN="{h_align}">'
            f'{text}</OZTABLELABEL>'
        )
    return "\\r\\n".join(xml)

def _build_page_header_band(header_config: dict, content_width: float) -> tuple:
    height = round(header_config.get("height", 50), 3)
    labels = header_config.get("labels", [])
    
    labels_xml = _build_static_table_labels(labels, "FixedTableLabel")
    count = len(labels)
    
    band = (
        f'<OZBAND NAME="PageHeaderBand1" WIDTH="{content_width}" '
        f'HEIGHT="{height}" BANDTYPE="1" COUNT="{count}">\\r\\n'
        f'<OZTABLESTATIC NAME="FixedTable1" LEFT="0" TOP="0" '
        f'WIDTH="{content_width}" HEIGHT="{height}">\\r\\n'
        f'{labels_xml}\\r\\n'
        f'</OZTABLESTATIC>\\r\\n'
        f'</OZBAND>'
    )
    return band, height

def _build_data_band_with_table(body_config: dict, content_width: float, odi_name: str, band_top: float) -> tuple:
    height = round(body_config.get("height", 40), 3)
    title_labels = body_config.get("title_labels", [])
    data_labels = body_config.get("data_labels", [])
    
    # 엑셀 전체 폭이 OZ용지 폭과 다를 수 있으므로 폭을 content_width에 맞추거나 그대로 둡니다.
    # 여기서는 추출된 정확한 좌표 그대로 사용합니다.
    table_width = content_width
    
    titles_xml = []
    for idx, lbl in enumerate(title_labels):
        text = escape_xml(lbl.get("text", ""))
        left = round(lbl.get("left", 0), 3)
        top = round(lbl.get("top", 0), 3)
        width = round(lbl.get("width", 100), 3)
        h = round(lbl.get("height", 20), 3)
        h_align = lbl.get("h_align", "0")
        
        border_attrs = 'DRAWLEFT="1" DRAWTOP="1" DRAWRIGHT="1" DRAWBOTTOM="1" ' if lbl.get("has_border") else 'DRAWLEFT="0" DRAWTOP="0" DRAWRIGHT="0" DRAWBOTTOM="0" '
        bg_color_attr = f'BGCOLOR="-5186329" ' if lbl.get("bg_color") != "none" else ""
        
        titles_xml.append(
            f'<OZTTLABEL NAME="TableTitle{idx + 1}" LEFT="{left}" TOP="{top}" '
            f'WIDTH="{width}" HEIGHT="{h}" '
            f'{bg_color_attr}{border_attrs}'
            f'DASHOFFSETLEFT="0" DASHOFFSETTOP="0" DASHOFFSETRIGHT="0" DASHOFFSETBOTTOM="0" '
            f'HALIGN="{h_align}" TABLEINDEX="{idx}">{text}</OZTTLABEL>'
        )
        
    values_xml = []
    field_items = []
    sell_labels = []
    datasets = set()
    
    for idx, lbl in enumerate(data_labels):
        text = lbl.get("text", "")
        left = round(lbl.get("left", 0), 3)
        top = round(lbl.get("top", 0), 3)
        width = round(lbl.get("width", 100), 3)
        h = round(lbl.get("height", 20), 3)
        h_align = lbl.get("h_align", "0")
        
        border_attrs = 'DRAWLEFT="1" DRAWTOP="1" DRAWRIGHT="1" DRAWBOTTOM="1" ' if lbl.get("has_border") else 'DRAWLEFT="0" DRAWTOP="0" DRAWRIGHT="0" DRAWBOTTOM="0" '
        
        # <DATASET:COLNAME> 파싱
        dataset_name = "TR_VIEW"
        col_name = f"COL{idx+1}"
        m = re.search(r'<([^:]+):([^>]+)>', text)
        if m:
            dataset_name = m.group(1)
            col_name = m.group(2)
        
        datasets.add(dataset_name)
        field_items.append(str(idx))
        sell_labels.append(col_name)
        
        values_xml.append(
            f'<OZGROUPLABEL NAME="TableValue{idx + 1}" LEFT="{left}" '
            f'TOP="{top}" WIDTH="{width}" HEIGHT="{h}" '
            f'ODINAME="{odi_name}" DATASETNAME="{dataset_name}" '
            f'COLNAME="{col_name}" '
            f'{border_attrs}'
            f'DASHOFFSETLEFT="0" DASHOFFSETTOP="0" DASHOFFSETRIGHT="0" DASHOFFSETBOTTOM="0" '
            f'HALIGN="{h_align}" '
            f'AUTOSIZE="true" AUTOFONTSIZE="smallerOnly" NULLTYPE="5" '
            f'PRIORLABELNAME="Root" TABLEINDEX="{idx}"/>'
        )
        
    field_items_str = "#%$oz*&amp;^".join(field_items)
    sell_labels_str = "#%$oz*&amp;^".join(sell_labels)
    default_dataset = list(datasets)[0] if datasets else "TR_VIEW"
    
    table_xml = (
        f'<OZTABLE NAME="Table1" WIDTH="{table_width}" HEIGHT="{height}" '
        f'ODINAME="{odi_name}" DATASET="{default_dataset}" '
        f'HAVETITLE="true" '
        f'TABLEFIELDITEMS="{field_items_str}" '
        f'TABLETITLEITEMS="{field_items_str}" '
        f'SELLABEL="{sell_labels_str}">\\r\\n'
        f'{"\\r\\n".join(titles_xml)}\\r\\n'
        f'{"\\r\\n".join(values_xml)}\\r\\n'
        f'</OZTABLE>'
    )
    
    band = (
        f'<OZDATABAND NAME="DataBand1" LEFT="0" TOP="{band_top}" '
        f'WIDTH="{content_width}" HEIGHT="{height}" '
        f'ODINAME="{odi_name}" MASTER="Report1" BANDTYPE="4" COUNT="1" '
        f'HEADERDUMMY="" FOOTERDUMMY="" SUBDATALIST="" UMDFIELDLIST="" '
        f'DATASOURCENAME="{default_dataset}">\\r\\n'
        f'{table_xml}\\r\\n'
        f'</OZDATABAND>'
    )
    return band, height, list(datasets)

def _build_page_footer_band(footer_config: dict, content_width: float, content_height: float) -> tuple:
    height = round(footer_config.get("height", 30), 3)
    labels = footer_config.get("labels", [])
    footer_top = content_height - height
    
    labels_xml = _build_static_table_labels(labels, "FixedTableLabel")
    count = len(labels)
    
    band = (
        f'<OZBAND NAME="PageFooterBand1" LEFT="0" TOP="{footer_top}" '
        f'WIDTH="{content_width}" HEIGHT="{height}" BANDTYPE="9" COUNT="{count}">\\r\\n'
        f'<OZTABLESTATIC NAME="FixedTable2" LEFT="0" TOP="0" '
        f'WIDTH="{content_width}" HEIGHT="{height}">\\r\\n'
        f'{labels_xml}\\r\\n'
        f'</OZTABLESTATIC>\\r\\n'
        f'</OZBAND>'
    )
    return band, height

def _build_odi_list(odi_ref_name: str, odi_filename: str, dataset_names: list) -> str:
    formsets = "\\r\\n".join(
        f'<OZFORMSET NAME="{ds}"/>' for ds in dataset_names
    )
    param_formset = (
        '<OZFORMSET NAME="OZParam">\\r\\n'
        '<PARAMFIELD FIELDNAME="rObjectId" VALUE="090186a08052e65e"/>\\r\\n'
        '</OZFORMSET>'
    )
    return (
        f'<OZODILIST>\\r\\n'
        f'<OZODIITEM NAME="{odi_ref_name}" CATEGORY="pl/pla" ODINAME="{odi_filename}">\\r\\n'
        f'{formsets}\\r\\n'
        f'{param_formset}\\r\\n'
        f'</OZODIITEM>\\r\\n'
        f'<OZFORMPARAMS/>\\r\\n'
        f'<OZRESOURCES/>\\r\\n'
        f'</OZODILIST>'
    )

def generate_ozr_xml(config: dict) -> str:
    odi_ref_name = config.get("odi_ref_name", "REPORT_ODI")
    odi_filename = config.get("odi_filename", "report.odi")
    header_config = config.get("header", {})
    body_config = config.get("body", {})
    footer_config = config.get("footer", {})
    
    content_width = round(CONTENT_WIDTH, 3)
    content_height = round(CONTENT_HEIGHT, 3)
    
    bands = []
    
    header_band_xml, header_height = _build_page_header_band(header_config, content_width)
    bands.append(header_band_xml)
    
    current_top = header_height
    
    data_band_xml, data_height, body_datasets = _build_data_band_with_table(
        body_config, content_width, odi_ref_name, current_top
    )
    bands.append(data_band_xml)
    current_top += data_height
    
    footer_band_xml, footer_height = _build_page_footer_band(footer_config, content_width, content_height)
    bands.append(footer_band_xml)
    
    bands.append(
        f'<OZBACKBAND NAME="BackgroundBand1" WIDTH="{content_width}" '
        f'HEIGHT="{content_height}" BANDTYPE="101"/>'
    )
    bands.append(
        f'<OZFOREBAND NAME="ForegroundBand1" WIDTH="{content_width}" '
        f'HEIGHT="{content_height}" BANDTYPE="103"/>'
    )
    
    bands_xml = "\\r\\n".join(bands)
    
    dataset_names = set(body_datasets)
    dataset_names.add("TR_LIST")
    dataset_names.add("TR_VIEW")
    
    odi_list = _build_odi_list(odi_ref_name, odi_filename, list(dataset_names))
    
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>\\r\\n'
        f'<OZREPORT WIDTH="{content_width}" HEIGHT="{content_height}" '
        f'VIRTUALX="1.4" VIRTUALY="1.4" '
        f'MAPMODE="4" DisplayDPI="96" FontDPI="0" '
        f'REPORT_PREVIEW_OPTION="viewer.smartframesize=true" '
        f'REPORTTYPE="1" GRIDSIZE="0.567, 0.567" JSCRIPTRULE="1" '
        f'InputVertRule="true" InputVertAlign="true" '
        f'InputBoxMultiLineRule="true" InputTextExVertAlign="true" '
        f'SCREENTOOL="1" IGNORENULLDATA="false" '
        f'Language="ko/KR" LEFTMARGIN="{LEFT_MARGIN}" TOPMARGIN="{TOP_MARGIN}" '
        f'RIGHTMARGIN="{RIGHT_MARGIN}" BOTTOMMARGIN="{BOTTOM_MARGIN}" '
        f'BookBinding="false" Orientation="false" '
        f'PAPERWIDTH="{PAPER_WIDTH}" PAPERHEIGHT="{PAPER_HEIGHT}">\\r\\n'
        f'{_build_version()}\\r\\n'
        f'{_build_basic_label()}\\r\\n'
        f'{_build_default_label()}\\r\\n'
        f'<REPORTINFO NAME="Report1" ReportVirtualSize="{content_width}, {content_height}" '
        f'LEFTMARGIN="{LEFT_MARGIN}" TOPMARGIN="{TOP_MARGIN}" '
        f'RIGHTMARGIN="{RIGHT_MARGIN}" BOTTOMMARGIN="{BOTTOM_MARGIN}" '
        f'BOOKBINDING="0" SUBDATALIST="DataBand1" '
        f'ORIENTATION="false" '
        f'PAPERWIDTH="{PAPER_WIDTH}" PAPERHEIGHT="{PAPER_HEIGHT}" '
        f'DRAWLEFT="0" DRAWTOP="0" DRAWRIGHT="0" DRAWBOTTOM="0" '
        f'WIDTH="{content_width}" HEIGHT="{content_height}">\\r\\n'
        f'{bands_xml}\\r\\n'
        f'</REPORTINFO>\\r\\n'
        f'<OZPARAMETERTOOLBARS NAME="ParameterToolbar1"/>\\r\\n'
        f'<OZFONTDESC>\\r\\n<OZFONTFAMILY/>\\r\\n</OZFONTDESC>\\r\\n'
        f'{odi_list}\\r\\n'
        f'<OZGRIDINFO GRIDOPERATE="False" GRIDSHOW="False" '
        f'GRIDX="0.567" GRIDY="0.567" GRIDTYPE="1"/>\\r\\n'
        f'<OZFORMIDINFO/>\\r\\n'
        f'</OZREPORT>\\r\\n'
    )
    
    return xml

def build_ozr_file(xml_content: str) -> bytes:
    return OZR_HEADER + xml_content.encode('utf-8')

def generate_ozr(config: dict) -> bytes:
    xml = generate_ozr_xml(config)
    return build_ozr_file(xml)
'''

with open(r'd:\DEV\oz\backend\xlsx_parser.py', 'w', encoding='utf-8') as f:
    f.write(xlsx_parser_code)
    
with open(r'd:\DEV\oz\backend\ozr_generator.py', 'w', encoding='utf-8') as f:
    f.write(ozr_generator_code)

print("Parser and Generator updated successfully.")
