"""
XLSX 파서 모듈 (개선판)
병합 셀(Merged Cells)과 데이터 바인딩(<DATASET:COLNAME>) 패턴을 완벽히 인식하여 레이아웃을 정확히 추출합니다.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import io
import re

EXCEL_COL_WIDTH_TO_OZ = 7.0
EXCEL_ROW_HEIGHT_TO_OZ = 1.33

def parse_xlsx(filepath) -> dict:
    if isinstance(filepath, bytes):
        file_bytes = io.BytesIO(filepath)
    else:
        with open(filepath, "rb") as f:
            file_bytes = io.BytesIO(f.read())
            
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    ws = wb.active
    
    # 1. 실제 사용된 최대 컬럼 찾기 (값이 있거나 테두리가 있는 컬럼)
    max_used_col = 1
    max_row = ws.max_row or 1
    max_col_candidate = ws.max_column or 1
    for r in range(1, max_row + 1):
        for c in range(1, max_col_candidate + 1):
            cell = ws.cell(row=r, column=c)
            has_val = cell.value is not None and str(cell.value).strip() != ""
            has_border = False
            if cell.border:
                if cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style:
                    has_border = True
            if has_val or has_border:
                if c > max_used_col:
                    max_used_col = c

    # 2. 사용된 영역 기준 스케일 계수(Scale Factor) 계산
    # 목표 가로 사이즈: 806.018 (A4 가로 842 - 좌우 여백)
    raw_col_widths = []
    for c in range(1, max_used_col + 1):
        col_letter = get_column_letter(c)
        dim = ws.column_dimensions.get(col_letter)
        raw_col_widths.append(dim.width if dim and dim.width else 8.43)
        
    raw_total_width = sum(raw_col_widths)
    TARGET_WIDTH = 806.018
    scale_factor = TARGET_WIDTH / raw_total_width if raw_total_width > 0 else 1.0

    # 3. 컬럼 폭 스케일링 적용 (빈 컬럼 포함 전체)
    col_widths = []
    for c in range(1, max_col_candidate + 1):
        col_letter = get_column_letter(c)
        dim = ws.column_dimensions.get(col_letter)
        width = dim.width if dim and dim.width else 8.43
        col_widths.append(width * scale_factor)
        
    # 누적 X 좌표 (컬럼별)
    col_x_offsets = [0]
    for w in col_widths:
        col_x_offsets.append(col_x_offsets[-1] + w)
        
    # 4. 행 높이 스케일링 적용 (비율 고정)
    Y_SCALE_FACTOR = 1.0
    row_heights = []
    for r in range(1, max_row + 1):
        rd = ws.row_dimensions.get(r)
        height = rd.height if rd and rd.height else 15.0
        row_heights.append(height * Y_SCALE_FACTOR)
        
    row_y_offsets = [0]
    for h in row_heights:
        row_y_offsets.append(row_y_offsets[-1] + h)
        
    # 병합 셀 맵핑 (Merged Cell Range 찾기)
    merged_ranges = ws.merged_cells.ranges
    merged_map = {} # (r, c) -> range
    for mr in merged_ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_map[(r, c)] = mr
                
    cells_data = []
    
    for r in range(1, max_row + 1):
        for c in range(1, max_col_candidate + 1):
            is_vertically_merged = False
            if (r, c) in merged_map:
                mr = merged_map[(r, c)]
                # 병합된 영역의 Top-Left 셀이 아니면 무시 (그리기 중복 방지)
                if r != mr.min_row or c != mr.min_col:
                    continue
                
                # 너비와 높이 계산
                left = col_x_offsets[mr.min_col - 1]
                top = row_y_offsets[mr.min_row - 1]
                width = col_x_offsets[mr.max_col] - left
                height = row_y_offsets[mr.max_row] - top
                
                if mr.max_row > mr.min_row and mr.max_col == mr.min_col:
                    is_vertically_merged = True
            else:
                left = col_x_offsets[c - 1]
                top = row_y_offsets[r - 1]
                width = col_x_offsets[c] - left
                height = row_y_offsets[r] - top
                
            cell = ws.cell(row=r, column=c)
            
            # 값이 있거나 테두리가 있거나 배경색이 있으면 유효한 셀로 간주
            val = cell.value
            val_str = str(val).strip() if val is not None else ""
            has_border = _has_border(cell)
            bg_color = _get_fill_color(cell)
            
            if not val_str and not has_border and bg_color == "none":
                continue
                
            cells_data.append({
                "row": r,
                "col": c,
                "text": val_str,
                "left": left,
                "top": top,
                "width": width,
                "height": height,
                "bold": cell.font.bold if cell.font else False,
                "font_size": cell.font.size if cell.font and cell.font.size else 10,
                "h_align": _map_h_align(cell.alignment.horizontal if cell.alignment else "left"),
                "v_align": _map_v_align(cell.alignment.vertical if cell.alignment else "center"),
                "has_border": has_border,
                "bg_color": bg_color,
                "vertically_merged": is_vertically_merged
            })
            
    # 영역 분류 (Data Row 찾기: <DATASET:COLNAME> 패턴이 있는 행)
    data_row_num = -1
    for cell in cells_data:
        if re.search(r'<[^:]+:[^>]+>', cell["text"]):
            data_row_num = cell["row"]
            break
            
    if data_row_num == -1:
        # 패턴이 없으면 기존 로직처럼 대략적으로 중간을 바디로 잡거나, 전체를 헤더로
        data_row_num = max_row // 2 if max_row > 2 else 2
        
    # 데이터 행의 셀 너비 경계선(X 좌표) 추출
    data_x_bounds = set()
    for c in cells_data:
        if c["row"] == data_row_num:
            data_x_bounds.add(round(c["left"], 1))
            data_x_bounds.add(round(c["left"] + c["width"], 1))
            
    # [NEW] 엑셀 12~30의 연속된 형태처럼, DataRow와 동일한 그리드(X좌표)를 가지는 반복행들 중
    # <DATASET> 태그가 없는 행들은 렌더링 대상에서 제외(1칸만 남기기 위함)
    repeating_rows = set()
    for r in range(1, max_row + 1):
        if r == data_row_num:
            continue
        row_cells = [c for c in cells_data if c["row"] == r]
        if not row_cells:
            continue
            
        row_x_bounds = set()
        for c in row_cells:
            row_x_bounds.add(round(c["left"], 1))
            row_x_bounds.add(round(c["left"] + c["width"], 1))
            
        if data_x_bounds:
            matching_bounds = row_x_bounds.intersection(data_x_bounds)
            match_ratio = len(matching_bounds) / len(data_x_bounds) if len(data_x_bounds) > 0 else 0
            # DataRow와 거의 완벽하게(90% 이상) 그리드가 일치하면 반복행으로 간주 (헤더나 푸터는 병합으로 인해 일치율이 보통 다름)
            if match_ratio >= 0.9:
                repeating_rows.add(r)
                
    # cells_data에서 반복행 제거
    cells_data = [c for c in cells_data if c["row"] not in repeating_rows]
        
    # 데이터 행의 셀 너비 경계선(X 좌표) 추출
    data_x_bounds = set()
    for c in cells_data:
        if c["row"] == data_row_num:
            data_x_bounds.add(round(c["left"], 1))
            data_x_bounds.add(round(c["left"] + c["width"], 1))
            
    # 데이터 행(data_row_num) 위쪽으로 스캔하여 Table Header 영역 찾기
    # 1) Dataset 바로 위의 빈 행들 무시
    current_r = data_row_num - 1
    while current_r > 0:
        if current_r in repeating_rows:
            current_r -= 1
            continue
        row_cells = [c for c in cells_data if c["row"] == current_r]
        if not row_cells:
            current_r -= 1
            continue
        break
        
    table_header_start = current_r + 1 if current_r < data_row_num - 1 else data_row_num
    
    for r in range(current_r, 0, -1):
        if r in repeating_rows:
            continue
            
        row_cells = [c for c in cells_data if c["row"] == r]
        if not row_cells:
            break
            
        has_borders = any(c["has_border"] for c in row_cells)
        if not has_borders:
            break
            
        # 해당 행의 x 경계선 수집
        row_x_bounds = set()
        for c in row_cells:
            row_x_bounds.add(round(c["left"], 1))
            row_x_bounds.add(round(c["left"] + c["width"], 1))
            
        # 데이터 행의 X 경계선과 얼마나 일치하는지 비율 계산
        if data_x_bounds:
            matching_bounds = row_x_bounds.intersection(data_x_bounds)
            match_ratio = len(matching_bounds) / len(data_x_bounds)
        else:
            match_ratio = 1.0
            
        # 일치율이 25% 미만이면 Page Header 영역이라고 간주
        if r < current_r and match_ratio < 0.25:
            break
            
        table_header_start = r
            
    footer_start = data_row_num + 1
    for r in range(data_row_num + 1, max_row + 1):
        row_cells = [c for c in cells_data if c["row"] == r]
        if not row_cells:
            continue
            
        has_borders = any(c["has_border"] for c in row_cells)
        if not has_borders:
            footer_start = r
            break
            
        row_x_bounds = set()
        for c in row_cells:
            row_x_bounds.add(round(c["left"], 1))
            row_x_bounds.add(round(c["left"] + c["width"], 1))
            
        if data_x_bounds:
            matching_bounds = row_x_bounds.intersection(data_x_bounds)
            match_ratio = len(matching_bounds) / len(data_x_bounds)
        else:
            match_ratio = 1.0
            
        # 데이터 그리드(가로 폭 분할)가 깨지면 거기가 푸터의 시작점
        if match_ratio < 0.5:
            footer_start = r
            break

    header_cells = []
    table_header_cells = []
    table_data_cells = []
    footer_cells = []
    
    for c in cells_data:
        r = c["row"]
        if r < table_header_start:
            header_cells.append(c)
        elif table_header_start <= r <= current_r:
            table_header_cells.append(c)
        elif current_r < r < data_row_num:
            pass # Ignore gap rows
        elif r == data_row_num:
            table_data_cells.append(c)
        elif r >= footer_start:
            footer_cells.append(c)
            
    # 높이 계산
    header_height = row_y_offsets[table_header_start - 1] if table_header_start > 1 else 0
    
    # 풋터 좌표 보정 (풋터는 DataBand (및 무시된 반복 데이터행들) 바로 밑에서 시작하므로 top을 0부터 다시 시작하게끔 보정)
    footer_start_y = row_y_offsets[footer_start - 1]
    for c in footer_cells:
        c["top"] -= footer_start_y
    footer_height = row_y_offsets[-1] - footer_start_y if footer_start_y < row_y_offsets[-1] else 20
    
    # DataBand 테이블의 Y 좌표 보정 (Table Header부터 0)
    table_start_y = row_y_offsets[table_header_start - 1]
    max_header_bottom = 0.0
    for c in table_header_cells:
        c["top"] -= table_start_y
        bottom = c["top"] + c["height"]
        if bottom > max_header_bottom:
            max_header_bottom = bottom
            
    # 데이터 행을 헤더 타이틀 바로 아래(max_header_bottom)에 밀착 배치 (R25 사양)
    data_row_h = table_data_cells[0]["height"] if table_data_cells else 20.0
    for c in table_data_cells:
        c["top"] = max_header_bottom
        
    body_height = max_header_bottom + data_row_h
    
    result = {
        "sheet_name": ws.title,
        "header": {
            "height": max(header_height, 20),
            "labels": header_cells
        },
        "body": {
            "title_labels": table_header_cells,
            "data_labels": table_data_cells,
            "height": max(body_height, 40)
        },
        "footer": {
            "height": max(footer_height, 20),
            "labels": footer_cells
        },
        "debug_info": {
            "table_header_start": table_header_start,
            "data_row_num": data_row_num,
            "footer_start": footer_start,
            "max_row": max_row
        },
        "total_width": col_x_offsets[-1]
    }
    
    wb.close()
    return result

def _get_fill_color(cell) -> str:
    try:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb != "00000000":
            return str(fill.fgColor.rgb)
    except:
        pass
    return "none"

def _has_border(cell) -> bool:
    try:
        border = cell.border
        if border:
            for side in [border.left, border.right, border.top, border.bottom]:
                if side and side.style and side.style != "none":
                    return True
    except:
        pass
    return False

def _map_h_align(align_str: str) -> str:
    # 0=left, 1=center, 2=right
    mapping = {"left": "0", "center": "1", "right": "2", "general": "0"}
    return mapping.get(str(align_str).lower(), "0")

def _map_v_align(align_str: str) -> str:
    # 0=top, 1=center, 2=bottom
    mapping = {"top": "0", "center": "1", "middle": "1", "bottom": "2"}
    return mapping.get(str(align_str).lower(), "1")

def get_xlsx_summary(parsed: dict) -> dict:
    """파싱 결과의 요약 정보를 반환합니다."""
    return {
        "sheet_name": parsed.get("sheet_name", ""),
        "header_labels_count": len(parsed["header"]["labels"]),
        "header_height": parsed["header"]["height"],
        "body_title_labels_count": len(parsed["body"]["title_labels"]),
        "body_data_labels_count": len(parsed["body"]["data_labels"]),
        "body_height": parsed["body"]["height"],
        "footer_labels_count": len(parsed["footer"]["labels"]),
        "footer_height": parsed["footer"]["height"],
        "total_width": parsed.get("total_width", 0),
    }
